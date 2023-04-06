import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
def registeropencart(firstname,lastname,email,password):
    driver=webdriver.Chrome(ChromeDriverManager().install())
    driver.get("https://demo.opencart.com/")
    driver.find_element(By.LINK_TEXT, "My Account").click()
    driver.find_element(By.LINK_TEXT, 'Register').click()
    time.sleep(5)
    print(driver.title)
    driver.find_element(By.ID, "input-firstname").send_keys(firstname)
    driver.find_element(By.ID, "input-lastname").send_keys(lastname)
    driver.find_element(By.ID, "input-email").send_keys(email)
    driver.find_element(By.ID, "input-password").send_keys(password)
    driver.find_element(By.NAME,"agree").click()
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    time.sleep(3)
    driver.quit()

excelpath = "C:\R\demo1py.xlsx"
workbook = openpyxl.load_workbook(excelpath)
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column
for i in range(2, rows + 1):
    fname = sheet.cell(row=i, column=1).value
    lname = sheet.cell(row=i, column=2).value
    eml = sheet.cell(row=i, column=3).value
    pword = sheet.cell(row=i, column=4).value
    registeropencart(fname,lname,eml,pword)