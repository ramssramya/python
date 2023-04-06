import time
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import pytest
from opencart.pageobjectregister import registerpage
@pytest.fixture()
def setup():
    print("run before every method")
def test_def1(setup):
    print("hello world")
def test_registeropencart1(setup):
    excelpath = "C:\R\demo1py.xlsx"
    workbook = openpyxl.load_workbook(excelpath)
    sheet = workbook.active
    rows = sheet.max_row
    cols = sheet.max_column
    print(rows)
    for i in range(2, rows + 1):
        print(i)
        fname = sheet.cell(row=i, column=1).value
        lname = sheet.cell(row=i, column=2).value
        eml = sheet.cell(row=i, column=3).value
        pword = sheet.cell(row=i, column=4).value
        driver = webdriver.Chrome(ChromeDriverManager().install())
        driver.get("https://demo.opencart.com/")
        rp = registerpage(driver)
        rp.clickonmyaccount()
        rp.clickonregisterlink()
        time.sleep(5)
        rp.enterfirstname(fname)
        rp.enterlastname(lname)
        time.sleep(4)
        rp.enteremail(eml)
        rp.enterpassword(pword)
        time.sleep(5)
        rp.clickonname()
        rp.clickonsubmit()
        time.sleep(5)
