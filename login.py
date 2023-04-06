import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import pytest
class testcase_001:
  def loginopencart(username, password):
    driver=webdriver.Chrome(ChromeDriverManager().install())
    driver.get("https://demo.opencart.com/")
    driver.find_element(By.LINK_TEXT,"My Account").click()
    driver.find_element(By.LINK_TEXT,"Login").click()
    time.sleep(5)
    if driver.title=="Your Store":
      print("title is matched")
    else:
      print("title not matched")
    time.sleep(4)
    driver.find_element(By.ID,"input-email").send_keys(username)
    driver.find_element(By.ID,"input-password").send_keys(password)
    driver.find_element(By.XPATH,"//button[@type='submit']").click()
    time.sleep(5)
    driver.quit()

  excelpath="C:\R\pydemo.xlsx"
  workbook=openpyxl.load_workbook(excelpath)
  sheet=workbook.active
  rows=sheet.max_row
  cols=sheet.max_column
  for i in range(2,rows+1):
    print(i)
    uname=sheet.cell(row=i,column=1).value
    pword=sheet.cell(row=i,column=2).value
    loginopencart(uname,pword)