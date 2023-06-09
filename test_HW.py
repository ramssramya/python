import time
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import pytest
from opencart.pageobjects import loginpage
@pytest.fixture()
def setup():
    print("run before every method")
def test_def1(setup):
    print("hello world")
def test_loginopencart1(setup):
    excelpath="C:\R\pydemo.xlsx"
    workbook=openpyxl.load_workbook(excelpath)
    sheet=workbook.active
    rows=sheet.max_row
    cols=sheet.max_column
    print(rows)
    for i in range(2,rows+1):
        print(i)
        uname=sheet.cell(row=i,column=1).value
        pword=sheet.cell(row=i,column=2).value
        driver = webdriver.Chrome(ChromeDriverManager().install())
        driver.get("https://demo.opencart.com/")
        lp=loginpage(driver)
        lp.clickonmyaccount()
        lp.clickonloginlink()
        time.sleep(5)
        lp.enteremail(uname)
        lp.enterpassword(pword)
        lp.clickonsubmit()
        time.sleep(5)
