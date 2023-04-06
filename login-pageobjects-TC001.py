import time
import openpyxl
import pageobject as pageobject
from selenium import webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import pytest
class testcase_001:
  def loginopencart(self,username, password):
    excelpath="C:\R\pydemo.xlsx"
    workbook=openpyxl.load_workbook(excelpath)
    sheet=workbook.active
    rows=sheet.max_row
    cols=sheet.max_column
    print(rows)
    for i in range(2,rows+1):
        print(i)
        uname=sheet.cell(row=i,column=1).value
        pword=sheet.cell(row=i,column=2).value
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.get("https://demo.opencart.com/")
        self.lp=pageobject.loginpage(self.driver)
        pageobject.lp.clickonmyaccount(self)
    loginopencart(uname,pword)